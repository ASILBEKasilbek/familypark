# utils/misc.py
import qrcode
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment,PatternFill
from datetime import datetime
import os
from models import *
from sqlalchemy import select

def generate_qr(link: str, filename: str = "qr_temp.png"):
    qr = qrcode.QRCode(box_size=12, border=5)
    qr.add_data(link)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1a1a1a", back_color="white")
    img.save(filename)
    return filename

def create_excel(users, headers=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "FamilyPark "

    if headers is None:
        headers = ["#", "Telegram ID", "Ism", "Username", "Telefon", "Source", "Ro'yxatdan o'tgan vaqti"]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")  
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, user in enumerate(users, start=1):
        row = [
            idx,
            user.telegram_id,
            user.first_name or "-",
            f"@{user.username}" if user.username else "-",
            user.phone,
            user.source,
            user.registered_at.strftime("%d.%m.%Y %H:%M")
        ]
        ws.append(row)

        cell = ws.cell(row=idx + 1, column=8)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    filename = f"FamilyPark_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    wb.save(filename)
    return filename


async def export_attendance_excel(async_session):
    async with async_session() as session:
        result = await session.execute(
            select(
                User.first_name,
                User.phone,
                AttendanceLog.place,
                AttendanceLog.marked_at
            ).join(AttendanceLog, AttendanceLog.user_id == User.telegram_id)
        )
        data = result.all()  
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["№", "Ism", "Telefon", "Joy", "Kelgan vaqti"]

    bold_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, row_data in enumerate(data, start=2):
        first_name, phone, place, marked_at = row_data
        ws.cell(row=row_num, column=1, value=row_num-1)
        ws.cell(row=row_num, column=2, value=first_name)
        ws.cell(row=row_num, column=3, value=phone)
        ws.cell(row=row_num, column=4, value=place)
        ws.cell(row=row_num, column=5, value=marked_at.strftime("%Y-%m-%d %H:%M"))

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 3

    file_name = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = f"/tmp/{file_name}"
    wb.save(file_path)

    return file_path
